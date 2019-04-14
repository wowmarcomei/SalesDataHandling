from openpyxl import load_workbook
import re
import time,datetime

# 输入文件
_BASE_FILE_NAME_ = '1-base.xlsx'
_ORDER_LIST_FILE_NAME_ = '2-orders-list.xlsm'
_ORDER_DETAILS_FILE_NAME_ = '3-orders-details.xlsx'
_REVENUE_FILE_NAME_ = '4-revenueCn.xlsx'
# 输出文件
_FIRST_TEMP_OUTPUT_FILE_NAME_ = 'temp_base_1.xlsx'
_SECOND_TEMP_OUTPUT_FILE_NAME_ = 'temp_base_2.xlsx'
_FINAL_OUTPUT_FILE_NAME_ = '最终输出件.xlsx'

# 打印LOG文件
LOG_FILE = open('./log.txt','w+')

# 调试模式
_DEBUG_ENABLE_ = False

# 读excel文件,指定参数data_only=True,拷贝源excel数据时不会拷贝源中的公式,默认data_only=False
print('*'*100+'\n')
print('正在加载《{0}》文件与《{1}》文件,文件较大请稍等...'.format(_BASE_FILE_NAME_,_ORDER_LIST_FILE_NAME_))
base_file = load_workbook(filename = _BASE_FILE_NAME_)
orders_file = load_workbook(filename = _ORDER_LIST_FILE_NAME_)  

def get_sales_mode(name):
        # unicode编码
        # 1-认证渠道项目销售  --> \u8ba4\u8bc1\u6e20\u9053\u9879\u76ee\u9500\u552e
        # 2-运营商转售 --> \u8fd0\u8425\u5546\u8f6c\u552e
        # 3-认证渠道分销  --> \u8ba4\u8bc1\u6e20\u9053\u5206\u9500
        # 4-直销  --> \u76f4\u9500
        if re.match(u'^\u8ba4\u8bc1\u6e20\u9053',name):
                return('渠道')
        elif re.match(u'^\u8fd0\u8425\u5546\u8f6c\u552e',name):
                return('运营商')
        elif re.match(u'^\u76f4\u9500',name):
                return('直销')
        else:
                return('NA')            

def first_proceed_base(base_file = base_file, input_file = orders_file):
    print('*'*100+'\n')
    print('正在基于《{0}》文件整理《{1}》文件,请稍等...'.format(_ORDER_LIST_FILE_NAME_,_BASE_FILE_NAME_))
    LOG_FILE.write('正在基于《{0}》文件整理《{1}》文件,请稍等...'.format(_ORDER_LIST_FILE_NAME_,_BASE_FILE_NAME_))
    # 获取到两个文件的打开时的那个活动sheet，其中第一个为"非服设备存量",第二个为“order list”
    base_file_sheet = base_file.get_active_sheet()
    input_file_sheet = input_file.get_active_sheet()

    # 获取base文件与orders文件最大行数
    base_rows = base_file_sheet.max_row  
    order_rows = input_file_sheet.max_row
    # print(base_rows,order_rows)

    # 根据合同号定义字典，dict1的键值为：base文件的合同号与行号，dict2的键值为：base文件的合同号与行号
    dict1 = {}
    dict2 = {}
    num = 0
    # 对dict赋值，从A2单元格开始算起，dict['x'] = y, x为合同号，y为行数
    for i in range(2,base_rows+1):
        dict1['{}'.format(base_file_sheet.cell(row=i,column=1).value)] = i
    for i in range(2,order_rows+1):
        dict2['{}'.format(input_file_sheet.cell(row=i,column=1).value)] = i    
    # dict2为订单列表中的合同号与行数组成key-value的字典，需要迭代查询是否在dict1中存在该键值
    for key2,value2 in dict2.items():
        if dict1.__contains__(key2):
            if _DEBUG_ENABLE_:
                print('附件2订单列表中的_A{0}_行的合同号_{1}_在基表中已经存在，不用关注.'.format(value2,key2))
                LOG_FILE.write('附件2订单列表中的_A{0}_行的合同号_{1}_在基表中已经存在，不用关注.'.format(value2,key2))
        else:
            if _DEBUG_ENABLE_:
                LOG_FILE.write('附件2订单列表中的_A{0}_行的合同号_{1}_在基表中不存在，将会添加该合同相关信息到基表的第_{2}_行'.format(value2,key2,base_rows+num)) 
                print('附件2订单列表中的_A{0}_行的合同号_{1}_在基表中不存在，将会添加该合同相关信息到基表的第_{2}_行'.format(value2,key2,base_rows+num)) 
            num = num + 1                            
            # 临时写入基线文件的A列x行(x即为迭代的行数)，值为订单列表input_file_sheet对应的单元格的值
            base_file_sheet['A{0}'.format(base_rows+num)] = input_file_sheet.cell(row=value2,column=1).value
            base_file_sheet['B{0}'.format(base_rows+num)] = input_file_sheet.cell(row=value2,column=2).value
            # 将订单列表中AE列的y行(y为迭代行数)中的时间提取出日期
            base_file_sheet['E{0}'.format(base_rows+num)] = re.findall(r'\d{4}\/\d{2}\/\d{2}',time.strftime("%Y/%m/%d %H:%M:%S", time.strptime(input_file_sheet['AE{}'.format(value2)].value, "%Y-%m-%d %H:%M:%S")))[0]
            base_file_sheet['G{0}'.format(base_rows+num)] = input_file_sheet.cell(row=value2,column=8).value
            # 调用函数get_sales_mode获取销售模式,因为re正则表达式只能匹配英文，所以自定义了一个函数
            base_file_sheet['J{0}'.format(base_rows+num)] = get_sales_mode(input_file_sheet.cell(row=value2,column=40).value)
            base_file_sheet['P{0}'.format(base_rows+num)] = input_file_sheet.cell(row=value2,column=43).value
            base_file_sheet['X{0}'.format(base_rows+num)] = input_file_sheet.cell(row=value2,column=39).value

            base_file_sheet['Z{0}'.format(base_rows+num)] = input_file_sheet.cell(row=value2,column=22).value
            base_file_sheet['AA{0}'.format(base_rows+num)] = input_file_sheet.cell(row=value2,column=13).value
            base_file_sheet['AB{0}'.format(base_rows+num)] = input_file_sheet.cell(row=value2,column=14).value
            base_file_sheet['AC{0}'.format(base_rows+num)] = input_file_sheet.cell(row=value2,column=16).value
            base_file_sheet['AD{0}'.format(base_rows+num)] = input_file_sheet.cell(row=value2,column=11).value
            base_file_sheet['AE{0}'.format(base_rows+num)] = input_file_sheet.cell(row=value2,column=4).value
            base_file_sheet['AF{0}'.format(base_rows+num)] = input_file_sheet.cell(row=value2,column=5).value

    # r_base_file已经修改，保存文件为《temp_base_1.xlsx》
    base_file.save(_FIRST_TEMP_OUTPUT_FILE_NAME_)
    return True

def second_proceed_base(base_file = base_file, input_file = orders_file):
    print('*'*100+'\n')
    print('正在基于《{0}》文件整理《{1}》文件,请稍等...'.format(_ORDER_DETAILS_FILE_NAME_,_FIRST_TEMP_OUTPUT_FILE_NAME_)) 
    LOG_FILE.write('正在基于《{0}》文件整理《{1}》文件,请稍等...'.format(_ORDER_DETAILS_FILE_NAME_,_FIRST_TEMP_OUTPUT_FILE_NAME_)) 
    # 获取到两个文件的打开时的那个活动sheet，其中第一个为"非服设备存量",第二个为"sheet 2"
    base_file_sheet = base_file.get_active_sheet()
    input_file_sheet = input_file.get_active_sheet()

    # 获取base文件与orders文件最大行数
    base_rows = base_file_sheet.max_row  
    order_rows = input_file_sheet.max_row
    # print(base_rows,order_rows)

    # 根据合同号定义字典，dict1的键值为：base文件的合同号与行号，dict2的键值为：base文件的合同号与行号
    dict1 = {}
    dict2 = {}
    num = 0
    # 对dict赋值，从A2单元格开始算起，dict['x'] = y, x为合同号，y为行数
    for i in range(2,base_rows+1):
        dict1['{}'.format(base_file_sheet.cell(row=i,column=1).value)] = i
    for i in range(2,order_rows+1):
        dict2['{}'.format(input_file_sheet.cell(row=i,column=1).value)] = i  

    # dict2为附件3详细订单中的合同号与行数组成key-value的字典，需要迭代查询是否在dict1中存在该键值
    for key2,value2 in dict2.items():
        if dict1.__contains__(key2):
            if _DEBUG_ENABLE_:
                print('附件3订单列表中的_A{0}_行的合同号_{1}_在基表中存在，将会添加该合同相关信息到基表的第_{2}_行'.format(value2,key2,dict1[key2]))
                LOG_FILE.write('附件3订单列表中的_A{0}_行的合同号_{1}_在基表中存在，将会添加该合同相关信息到基表的第_{2}_行'.format(value2,key2,dict1[key2]))
            # 临时写入基线文件的A列x行(x即为迭代的行数)，值为订单列表input_file_sheet对应的单元格的值
            base_file_sheet['AG{0}'.format(dict1[key2])] = input_file_sheet.cell(row=value2,column=2).value
            base_file_sheet['AH{0}'.format(dict1[key2])] = input_file_sheet.cell(row=value2,column=3).value
            base_file_sheet['AI{0}'.format(dict1[key2])] = input_file_sheet.cell(row=value2,column=4).value
            base_file_sheet['AJ{0}'.format(dict1[key2])] = input_file_sheet.cell(row=value2,column=5).value
            base_file_sheet['AK{0}'.format(dict1[key2])] = input_file_sheet.cell(row=value2,column=6).value
            base_file_sheet['AL{0}'.format(dict1[key2])] = input_file_sheet.cell(row=value2,column=7).value
            base_file_sheet['AM{0}'.format(dict1[key2])] = input_file_sheet.cell(row=value2,column=8).value
            base_file_sheet['AN{0}'.format(dict1[key2])] = input_file_sheet.cell(row=value2,column=9).value
            base_file_sheet['AO{0}'.format(dict1[key2])] = input_file_sheet.cell(row=value2,column=10).value
        else:
            if _DEBUG_ENABLE_:
                print('附件3订单列表中的_A{0}_行的合同号_{1}_在基表中已经不存在，不用关注.'.format(value2,key2))
                LOG_FILE.write('附件3订单列表中的_A{0}_行的合同号_{1}_在基表中已经不存在，不用关注.'.format(value2,key2))

    # temp_base_1已经修改，保存文件为《temp_base_2.xlsx》
    base_file.save(_SECOND_TEMP_OUTPUT_FILE_NAME_)
    return True

def third_proceed_base(base_file = base_file, input_file = orders_file):
    print('*'*100+'\n')
    print('正在基于《{0}》文件整理《{1}》文件,请稍等...'.format(_REVENUE_FILE_NAME_,_SECOND_TEMP_OUTPUT_FILE_NAME_))
    LOG_FILE.write('正在基于《{0}》文件整理《{1}》文件,请稍等...'.format(_REVENUE_FILE_NAME_,_SECOND_TEMP_OUTPUT_FILE_NAME_))
    # 获取到两个文件的打开时的那个活动sheet，其中第一个为"非服设备存量",第二个为"sheet 2"
    base_file_sheet = base_file.get_active_sheet()
    input_file_sheet = input_file.get_active_sheet()

    # 获取base文件与orders文件最大行数
    base_rows = base_file_sheet.max_row  
    order_rows = input_file_sheet.max_row
    # print(base_rows,order_rows)

    # 根据合同号定义字典，dict1的键值为：base文件的合同号与行号，dict2的键值为：base文件的合同号与行号
    dict1 = {}
    dict2 = {}
    num = 0
    # 对dict赋值，从A2单元格开始算起，dict['x'] = y, x为合同号，y为行数
    for i in range(2,base_rows+1):
        dict1['{}'.format(base_file_sheet.cell(row=i,column=1).value)] = i
    for i in range(2,order_rows+1):
        dict2['{}'.format(input_file_sheet.cell(row=i,column=1).value)] = i  

    # dict2为附件4详细订单中的合同号与行数组成key-value的字典，需要迭代查询是否在dict1中存在该键值
    for key2,value2 in dict2.items():
        if dict1.__contains__(key2):
            if _DEBUG_ENABLE_:
                print('附件4订单列表中的_A{0}_行的合同号_{1}_在基表中存在，将会添加该合同相关信息到基表的第_{2}_行'.format(value2,key2,dict1[key2]))
                LOG_FILE.write('附件4订单列表中的_A{0}_行的合同号_{1}_在基表中存在，将会添加该合同相关信息到基表的第_{2}_行'.format(value2,key2,dict1[key2]))
            # 临时写入基线文件的A列x行(x即为迭代的行数)，值为订单列表input_file_sheet对应的单元格的值
            # 由于读取的日期有的为空值，所以使用if判断语句赋值，只有读取日期不为空的时候才会进行正则表达式匹配
            base_file_sheet['K{0}'.format(dict1[key2])] = str(input_file_sheet.cell(row=value2,column=43).value.year)+'/'+str(input_file_sheet.cell(row=value2,column=43).value.month)+'/'+str(input_file_sheet.cell(row=value2,column=43).value.day) if input_file_sheet.cell(row=value2,column=43).value  else 'NA'
            # base_file_sheet['K{0}'.format(dict1[key2])] = input_file_sheet.cell(row=value2,column=43).value
            base_file_sheet['M{0}'.format(dict1[key2])] = input_file_sheet.cell(row=value2,column=38).value
            base_file_sheet['N{0}'.format(dict1[key2])] = input_file_sheet.cell(row=value2,column=52).value
        else:
            if _DEBUG_ENABLE_:
                print('附件4订单列表中的_A{0}_行的合同号_{1}_在基表中已经不存在，不用关注.'.format(value2,key2))
                LOG_FILE.write('附件4订单列表中的_A{0}_行的合同号_{1}_在基表中已经不存在，不用关注.'.format(value2,key2))                

    # temp_base_2已经修改，保存文件为《最终输出件.xlsx》
    base_file.save(_FINAL_OUTPUT_FILE_NAME_)
    return True

# 读excel文件,指定参数data_only=True,拷贝源excel数据时不会拷贝源中的公式,默认data_only=False
if first_proceed_base(base_file = base_file, input_file = orders_file):
    print('第1阶段整理完成，进入第2阶段')
    if second_proceed_base(load_workbook(filename = _FIRST_TEMP_OUTPUT_FILE_NAME_), load_workbook(filename = _ORDER_DETAILS_FILE_NAME_)):
        print('第2阶段整理完成，进入第3阶段')
        if third_proceed_base(load_workbook(filename = _SECOND_TEMP_OUTPUT_FILE_NAME_), load_workbook(filename = _REVENUE_FILE_NAME_,data_only=True)):
            print('恭喜，处理完成')
            LOG_FILE.close()
