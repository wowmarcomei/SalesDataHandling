from openpyxl import Workbook,load_workbook
import re
import time,datetime 

# 指定参数data_only=True,拷贝源excel数据时不会拷贝源中的公式
test_file = load_workbook(filename = '4-revenueCn.xlsx',data_only=True)
r_orders_file_sheet = test_file.get_active_sheet()

print(r_orders_file_sheet.cell(row=4120,column=52).value)

# 读excel文件
# r_base_file = load_workbook(filename = '1-base.xlsx')
# r_orders_file = load_workbook(filename = '2-orders-list.xlsm')
# r_orders_details_file = load_workbook(filename = '3-orders-details.xlsx')
# r_revenue_file = load_workbook(filename = '4-revenueCn.xlsx')

# 获取base文件与orders文件的相应sheet
# r_base_file_sheet = r_base_file['非服设备存量']
# r_orders_file_sheet = r_orders_file['order List'] 
# r_orders_details_file_sheet = r_orders_details_file['Sheet2']

# 获取base文件与orders文件最大行数
# base_rows = r_base_file_sheet.max_row  
# order_rows = r_orders_file_sheet.max_row
# details_rows = r_orders_details_file_sheet.max_row  

# dict1 = {}
# dict2 = {}
# num = 0

# 对dict赋值，从A2单元格开始算起，dict['x'] = y, x为合同号，y为行数
# for i in range(2,base_rows+1):
#     dict1['{}'.format(r_base_file_sheet.cell(row=i,column=1).value)] = i

# dict3 = {}
# for i in range(2,details_rows+1):
#     dict3['{}'.format(r_orders_details_file_sheet.cell(row=i,column=1).value)] = i    

# dict3为附件3详细订单中的合同号与行数组成key-value的字典，需要迭代查询是否在dict1中存在该键值
# for key3,value3 in dict3.items():
#         if dict1.__contains__(key3):
#                 print('附件3订单列表中的_A{0}_行的合同号_{1}_在基表中存在，将会添加该合同相关信息到基表的第_{2}_行'.format(value3,key3,dict1[key3]))
#         else:
#                 print('附件3订单列表中的_A{0}_行的合同号_{1}_在基表中已经不存在，不用关注.'.format(value3,key3))
